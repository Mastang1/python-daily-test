from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

report_path = "C:\\tangyapeng\\docs\\StarGather\\report_gen_proj\\mculong-test-ipcs_20251014_09_25_37.html"
case_in_redmine = [
    "IPCS_FAULT_TC_000",
    "IPCS_FAULT_TC_001",
    "IPCS_FAULT_TC_002",
    "IPCS_FAULT_TC_003",
    "IPCS_FAULT_TC_004",
    "IPCS_FAULT_TC_005",
    "IPCS_FAULT_TC_006",
    "IPCS_FAULT_TC_007",
    "IPCS_PERF_TC_000",
    "IPCS_PERF_TC_001",
    "IPCS_PERF_TC_002",
    "IPCS_PERF_TC_003",
    "IPCS_PERF_TC_004",
    "IPCS_PERF_TC_005",
    "IPCS_PERF_TC_006",
    "IPCS_TS_0_TC_000",
    "IPCS_TS_0_TC_001",
    "IPCS_TS_0_TC_002",
    "IPCS_TS_1_TC_000",
    "IPCS_TS_1_TC_001",
    "IPCS_TS_1_TC_002",
    "IPCS_TS_2_TC_000",
    "IPCS_TS_2_TC_001",
    "IPCS_TS_2_TC_002",
    "IPCS_TS_3_TC_000",
    "IPCS_TS_3_TC_001",
    "IPCS_TS_3_TC_002",
    "IPCS_TS_4_TC_000",
    "IPCS_TS_4_TC_001",
    "IPCS_TS_6_TC_000",
    "IPCS_TS_7_TC_000",
    "IPCS_TS_8_TC_000",
    "IPCS_TS_9_TC_000"
]

def search_between_strings(file_path, string_a, string_b):
    listTestInfo = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            
            found_a = False
            start_line = 0
            
            for line_num, line in enumerate(lines, 1):
                if string_a in line and not found_a:
                    found_a = True
                    start_line = line_num
                    # print(f"找到字符串A '{string_a}' 在行 {line_num}")
                    continue
                
                if found_a and string_b in line:
                    # print(f"找到字符串B '{string_b}' 在行 {line_num}")
                    # print(f"\n字符串A和字符串B之间的文本（行 {start_line} 到行 {line_num}）:")
                    # 打印从字符串A所在行到字符串B所在行之间的所有文本
                    for i in range(start_line - 1, line_num):
                        if "test begin >>>" in lines[i].strip():
                            for j in range(i - 1, line_num - 1):
                                listTestInfo.append(lines[j].strip())
                    break
                    
    except FileNotFoundError:
        print(f"文件 '{file_path}' 不存在")
    except Exception as e:
        print(f"发生错误: {e}")
    strTestInfo = "\n".join(listTestInfo)
    print(type(strTestInfo))
    return strTestInfo

def clean_string_for_excel(text):
    """
    清理字符串，使其适合写入Excel单元格
    """
    if text is None:
        return ""
    
    # 转换为字符串
    text = str(text)
    
    # 移除或替换可能引起问题的特殊字符
    # 移除控制字符（ASCII 0-31，除了制表符、换行符、回车符）
    text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)
    
    # 替换可能引起问题的字符
    text = text.replace('\x00', '')  # 空字符
    text = text.replace('\x01', '')  # 标题开始
    text = text.replace('\x02', '')  # 正文开始
    text = text.replace('\x03', '')  # 正文结束
    text = text.replace('\x04', '')  # 文件结束
    
    # 限制字符串长度（Excel单元格最大字符数为32767）
    if len(text) > 32000:
        text = text[:32000] + "...[内容被截断]"
    
    return text
def write_list_to_column(string_list, excel_file_path, column='P', start_row=2):
    """
    将字符串列表写入Excel文件的指定列
    
    参数:
    string_list: list - 要写入的字符串列表
    excel_file_path: str - Excel文件路径
    column: str - 列字母，默认为'P'
    start_row: int - 起始行号，默认为2（P2列）
    """
    try:
        # 加载工作簿
        wb = load_workbook(filename=excel_file_path)
        
        # 获取第一个工作表（或指定工作表）
        ws = wb["Test Case Report"]
        
        # 将字符串列表写入指定列
        for i, value in enumerate(string_list):
            cleaned_value = clean_string_for_excel(value)
            row_num = start_row + i
            cell_address = f"{column}{row_num}"
            ws[cell_address] = cleaned_value
        
        # 保存工作簿
        wb.save(excel_file_path)
        print(f"成功将{len(string_list)}个字符串写入到{excel_file_path}的{column}列")
        
    except Exception as e:
        print(f"写入Excel文件时出错: {e}")

def searchAllInList(strTcList=case_in_redmine):
    listTcInfos = []
    for strTc in strTcList:
        listTcInfos.append(search_between_strings(report_path, strTc, ">>> test end"))
    return listTcInfos

if __name__ == "__main__":

    # strTestInfo = search_between_strings(report_path, "IPCS_FAULT_TC_000", ">>> test end ")

    listTcInfos = searchAllInList()
    # for item in listTcInfos:
    #     print(item)
    
    listTest = ["1111", "2222", "3333"]
    write_list_to_column(listTcInfos, "C:\\tangyapeng\\docs\\StarGather\\report_gen_proj\\report-ipcs.xlsx")

