from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

report_path = "C:\\tangyapeng\\docs\\StarGather\\report_gen_proj\\mculong-test-ipcs_20251014_09_25_37.html"
case_in_redmine = {
"MPECAN_FW_TC_0002":"addr_SharedMemory_Access Passed",
"MPECAN_FW_TC_0003":"addr_Register_Access Passed",
"MPECAN_FW_TC_0160":"FIFO_Disable Test Passed",
"MPECAN_FW_TC_0160_1":"FIFO_Enable Test Passed",
"MPECAN_FW_TC_0161":"FIFO_PUSH Test Passed",
"MPECAN_FW_TC_0169":"FIFO_PUSH_WITH_MSB_CK Test Passed",
"MPECAN_FW_TC_0165":"FIFO_POP Test Passed",
"MPECAN_FW_TC_0167":"FIFO_POP_MULTIPLE_ENTRIES Test Passed",
"MPECAN_FW_TC_0166":"FIFO_POP_MULTIPLE_ENTRY_UNDERFLOW_CHECK Test Passed",
"MPECAN_FW_TC_0168":"FIFO_PUSH_SAME_ENTRY_CHECK Test Passed",
"MPECAN_FW_TC_0163":"FIFO_PUSH_15_ENTRY_CHK_FLG Test Passed",
"MPECAN_FW_TC_0164":"FIFO_PUSH_16_ENTRY_CHK_FULL Test Passed",
"MPECAN_FW_TC_0170":"FIFO_PUSH_15_ENTRY_CHK_FLG Test Failed",
"MPECAN_FW_TC_0137-1":"TXLUT_Disable Test Passed",
"MPECAN_FW_TC_0141":"TXLUT_Enable Test Passed",
"MPECAN_FW_TC_0138":"TXLUT_PUSH Test Passed",
"MPECAN_FW_TC_0142":"TXLUT_POP Test Passed",
"MPECAN_FW_TC_0139":"TXLUT_PUSH_MULTIPLE_ENTRY Test Passed",
"MPECAN_FW_TC_0140":"TXLUT_PUSH_MULTIPLE_ENTRY_FULL_CHECK Test Passed",
"MPECAN_FW_TC_0144":"TXLUT_POP_MULTIPLE_ENTRIES Test Passed",
"MPECAN_FW_TC_0143":"TXLUT_POP_MULTIPLE_ENTRY_UNDERFLOW_CHECK Test Passed",
"MPECAN_FW_TC_0145":"TXLUT_PUSH_REMOVE_CHECK Test Failed",
"MPECAN_FW_TC_0148":"TXLUT_PUSH_SAME_ENTRY_CHECK Test Passed",
"MPECAN_FW_TC_0147":"TXLUT_SEARCH_CHECK Test Passed",
"MPECAN_FW_TC_0158":"TXLUT_PUSH_WITH_MSB_CK Test Passed",
"MPECAN_FW_TC_0159":"TXLUT_POP_MULTIPLE_ENTRIES_CHK_PRIORITY Test Passed",
"MPECAN_FW_TC_0146":"TXLUT_POP_MULTIPLE_ENTRIES_CHK_LBEN Test Passed",
"MPECAN_FW_TC_0151":"Rxlut_can2can_filter_search",
"MPECAN_FW_TC_0152":"Rxlut_can2eth_filter_search",
"MPECAN_FW_TC_0153":"Rxlut_can2cm7_filter_search",
"MPECAN_FW_TC_0150":"Rxlut_Write_read_via_Addr Passed",
"MPECAN_FW_TC_0154":"Rxlut_read_write_Entry Passed",
"MPECAN_FW_TC_0085":"Set Filter with 4K Index Check Rxlut 0 Entry index Boundary Failed",
"MPECAN_FW_TC_0084":"Set Filter with 4K Index Check Rxlut 0 Remove Entry index Boundary Failed",
"MPECAN_FW_TC_0186":"Rxlut_read_write_Entry Failed",
"MPECAN_FW_TC_0065":"Rxlut_range_search Failed",
"MPECAN_FW_TC_0187":"Rxlut_range_search_With_Extern_ID Failed",
"MPECAN_FW_TC_0101":"Rxlut2 Function Read Entry Passed",
"MPECAN_FW_TC_0053":"Config Max Filter num Failed",
"MPECAN_FW_TC_0195":"LIN_Tx_Master_Mode Passed",
"MPECAN_FW_TC_0196":"LIN_Tx_Master_Mode_MUltiple_times Passed",
"MPECAN_FW_TC_0192":"LIN_FD_Tx_Rx_LoopBack_Mode Passed",
"MPECAN_FW_TC_0191":"LIN_Tx_Rx_LoopBack_Mode Passed",
"MPECAN_FW_TC_0194":"LIN_Tx_Rx_LoopBack_Intrl_Mode Passed",
"MPECAN_FW_TC_0193":"LIN_Tx_Rx_LoopBack_Intrl_Mode Passed",
"MPECAN_FW_TC_119":"CM7_Send_Resp_Chk_Lin_Drv_Slave Passed",
"MPECAN_FW_TC_0199":"LIN_SEND_ERR_IND_CHECK_FROM_CM7_Slave failed ,header indication not received",
"MPECAN_FW_TC_0124":"LIN_SEND_ERR_IND_CHECK_FROM_CM7_Slave failed ,header indication not received",
"MPECAN_FW_TC_0174":"LIN_Ctrl_Set_Boundrate Test Passed",
"MPECAN_FW_TC_0171":"LIN_Controller_Enable Lin Mode failed because getmode not correct",
"MPECAN_FW_TC_0172":"LIN_Controller_Disable Lin Mode failed because getmode not correct",
"MPECAN_FW_TC_0198":"LINFW_Tx_CM_Rx_LoopBack_Slave_Mode Passed",
"MPECAN_FW_TC_0197":"LINFW_Tx_CM7_Rx_LoopBack_Master_Mode Passed",
"MPECAN_FW_TC_0115":"LIN_Drop the first Tx when New Req MASTER Test Passed",
"MPECAN_FW_TC_0118":"4Channel LIN_Send_From_CM7_MASTER Test Failed",
"MPECAN_FW_TC_0066":"LIN_ADV_MASTER_LIN2ETH_STUB Test Failed",
"MPECAN_FW_TC_0093":"LIN_ADV_MASTER_ETH2LIN_STUB Test Passed",
"MPECAN_FW_TC_0076":"Check Lin Adv routing and filter tbl Failed",
"MPECAN_FW_TC_0080":"LIN_ADV_MASTER_LIN2ETH_STUB_Slave Test Failed",
"MPECAN_FW_TC_0087":"LIN_ADV_MASTER_ETH2LIN_STUB_Slave Test Failed",
"MPECAN_FW_TC_0087":"LIN_ADV_MASTER_ETH2LIN_STUB_Slave Test Failed",
"MPECAN_FW_TC_0240":"Sema_Lock Test Failed",
"MPECAN_FW_TC_0240":"Sema_Lock Test Failed",
"MPECAN_FW_TC_0056":"BCAN_CTRL_ERROR_NOTIF FAILED failed because Receive Error info is 0",
"MPECAN_FW_TC_0059":"BCAN_CTRL_ERROR_NOTIF FAILED failed because Receive Error info is 0",
"MPECAN_FW_TC_0028":"GET_FW_VERSION Passed. builf time is 2025-09-30 10",
"MPECAN_FW_TC_0189":"FAULTDETECT_CHECK_DRX_HEARTBEAT failed because timestamp value not correct",
"MPECAN_FW_TC_0107":"FAULTDETECT_CHECK_STATUS_AFTER_FAULT Passed",
"MPECAN_FW_TC_0098":"reload fw after reset Passed",
"MPECAN_FW_TC_0104":"check internal var after reset Passed",
"MPECAN_FW_TC_0111":"Gracefull reset check sent initPilatform cmd after reset failed because status value not correct",
"MPECAN_FW_TC_0112":"UnGracefull reset check sent initPilatform cmd after reset failed because status value not correct",
"MPECAN_FW_TC_0113":"reset with routing failed because status value not correct"
}
all_case_s = [
    "MPECAN_FW_TC_0000",
    "MPECAN_FW_TC_0001",
    "MPECAN_FW_TC_0002",
    "MPECAN_FW_TC_0003",
    "MPECAN_FW_TC_0004",
    "MPECAN_FW_TC_0005",
    "MPECAN_FW_TC_0006",
    "MPECAN_FW_TC_0007",
    "MPECAN_FW_TC_0008",
    "MPECAN_FW_TC_0009",
    "MPECAN_FW_TC_0010",
    "MPECAN_FW_TC_0011",
    "MPECAN_FW_TC_0012",
    "MPECAN_FW_TC_0013",
    "MPECAN_FW_TC_0014",
    "MPECAN_FW_TC_0015",
    "MPECAN_FW_TC_0016",
    "MPECAN_FW_TC_0017",
    "MPECAN_FW_TC_0018",
    "MPECAN_FW_TC_0019",
    "MPECAN_FW_TC_0020",
    "MPECAN_FW_TC_0021",
    "MPECAN_FW_TC_0022",
    "MPECAN_FW_TC_0023",
    "MPECAN_FW_TC_0024",
    "MPECAN_FW_TC_0025",
    "MPECAN_FW_TC_0026",
    "MPECAN_FW_TC_0027",
    "MPECAN_FW_TC_0028",
    "MPECAN_FW_TC_0029",
    "MPECAN_FW_TC_0030",
    "MPECAN_FW_TC_0031",
    "MPECAN_FW_TC_0032",
    "MPECAN_FW_TC_0033",
    "MPECAN_FW_TC_0034",
    "MPECAN_FW_TC_0035",
    "MPECAN_FW_TC_0036",
    "MPECAN_FW_TC_0037",
    "MPECAN_FW_TC_0038",
    "MPECAN_FW_TC_0039",
    "MPECAN_FW_TC_0040",
    "MPECAN_FW_TC_0041",
    "MPECAN_FW_TC_0042",
    "MPECAN_FW_TC_0043",
    "MPECAN_FW_TC_0044",
    "MPECAN_FW_TC_0045",
    "MPECAN_FW_TC_0046",
    "MPECAN_FW_TC_0047",
    "MPECAN_FW_TC_0048",
    "MPECAN_FW_TC_0049",
    "MPECAN_FW_TC_0050",
    "MPECAN_FW_TC_0051",
    "MPECAN_FW_TC_0052",
    "MPECAN_FW_TC_0053",
    "MPECAN_FW_TC_0054",
    "MPECAN_FW_TC_0055",
    "MPECAN_FW_TC_0056",
    "MPECAN_FW_TC_0057",
    "MPECAN_FW_TC_0058",
    "MPECAN_FW_TC_0059",
    "MPECAN_FW_TC_0060",
    "MPECAN_FW_TC_0061",
    "MPECAN_FW_TC_0062",
    "MPECAN_FW_TC_0063",
    "MPECAN_FW_TC_0064",
    "MPECAN_FW_TC_0065",
    "MPECAN_FW_TC_0066",
    "MPECAN_FW_TC_0067",
    "MPECAN_FW_TC_0068",
    "MPECAN_FW_TC_0069",
    "MPECAN_FW_TC_0070",
    "MPECAN_FW_TC_0071",
    "MPECAN_FW_TC_0072",
    "MPECAN_FW_TC_0073",
    "MPECAN_FW_TC_0074",
    "MPECAN_FW_TC_0075",
    "MPECAN_FW_TC_0076",
    "MPECAN_FW_TC_0077",
    "MPECAN_FW_TC_0078",
    "MPECAN_FW_TC_0079",
    "MPECAN_FW_TC_0080",
    "MPECAN_FW_TC_0081",
    "MPECAN_FW_TC_0082",
    "MPECAN_FW_TC_0083",
    "MPECAN_FW_TC_0084",
    "MPECAN_FW_TC_0085",
    "MPECAN_FW_TC_0086",
    "MPECAN_FW_TC_0087",
    "MPECAN_FW_TC_0088",
    "MPECAN_FW_TC_0089",
    "MPECAN_FW_TC_0090",
    "MPECAN_FW_TC_0091",
    "MPECAN_FW_TC_0092",
    "MPECAN_FW_TC_0093",
    "MPECAN_FW_TC_0094",
    "MPECAN_FW_TC_0095",
    "MPECAN_FW_TC_0096",
    "MPECAN_FW_TC_0097",
    "MPECAN_FW_TC_0098",
    "MPECAN_FW_TC_0099",
    "MPECAN_FW_TC_0100",
    "MPECAN_FW_TC_0101",
    "MPECAN_FW_TC_0102",
    "MPECAN_FW_TC_0103",
    "MPECAN_FW_TC_0104",
    "MPECAN_FW_TC_0105",
    "MPECAN_FW_TC_0106",
    "MPECAN_FW_TC_0107",
    "MPECAN_FW_TC_0108",
    "MPECAN_FW_TC_0109",
    "MPECAN_FW_TC_0110",
    "MPECAN_FW_TC_0111",
    "MPECAN_FW_TC_0112",
    "MPECAN_FW_TC_0113",
    "MPECAN_FW_TC_0114",
    "MPECAN_FW_TC_0115",
    "MPECAN_FW_TC_0116",
    "MPECAN_FW_TC_0117",
    "MPECAN_FW_TC_0119",
    "MPECAN_FW_TC_0120",
    "MPECAN_FW_TC_0121",
    "MPECAN_FW_TC_0122",
    "MPECAN_FW_TC_0123",
    "MPECAN_FW_TC_0124",
    "MPECAN_FW_TC_0125",
    "MPECAN_FW_TC_0126",
    "MPECAN_FW_TC_0127",
    "MPECAN_FW_TC_0128",
    "MPECAN_FW_TC_0129",
    "MPECAN_FW_TC_0130",
    "MPECAN_FW_TC_0131",
    "MPECAN_FW_TC_0132",
    "MPECAN_FW_TC_0133",
    "MPECAN_FW_TC_0134",
    "MPECAN_FW_TC_0135",
    "MPECAN_FW_TC_0136",
    "MPECAN_FW_TC_0137",
    "MPECAN_FW_TC_0138",
    "MPECAN_FW_TC_0139",
    "MPECAN_FW_TC_0140",
    "MPECAN_FW_TC_0141",
    "MPECAN_FW_TC_0142",
    "MPECAN_FW_TC_0143",
    "MPECAN_FW_TC_0144",
    "MPECAN_FW_TC_0145",
    "MPECAN_FW_TC_0146",
    "MPECAN_FW_TC_0147",
    "MPECAN_FW_TC_0148",
    "MPECAN_FW_TC_0149",
    "MPECAN_FW_TC_0150",
    "MPECAN_FW_TC_0151",
    "MPECAN_FW_TC_0152",
    "MPECAN_FW_TC_0153",
    "MPECAN_FW_TC_0154",
    "MPECAN_FW_TC_0158",
    "MPECAN_FW_TC_0159",
    "MPECAN_FW_TC_0160",
    "MPECAN_FW_TC_0161",
    "MPECAN_FW_TC_0162",
    "MPECAN_FW_TC_0163",
    "MPECAN_FW_TC_0164",
    "MPECAN_FW_TC_0165",
    "MPECAN_FW_TC_0166",
    "MPECAN_FW_TC_0167",
    "MPECAN_FW_TC_0168",
    "MPECAN_FW_TC_0169",
    "MPECAN_FW_TC_0170",
    "MPECAN_FW_TC_0171",
    "MPECAN_FW_TC_0172",
    "MPECAN_FW_TC_0173",
    "MPECAN_FW_TC_0174",
    "MPECAN_FW_TC_0175",
    "MPECAN_FW_TC_0176",
    "MPECAN_FW_TC_0177",
    "MPECAN_FW_TC_0178",
    "MPECAN_FW_TC_0179",
    "MPECAN_FW_TC_0180",
    "MPECAN_FW_TC_0181",
    "MPECAN_FW_TC_0182",
    "MPECAN_FW_TC_0183",
    "MPECAN_FW_TC_0184",
    "MPECAN_FW_TC_0185",
    "MPECAN_FW_TC_0186",
    "MPECAN_FW_TC_0187",
    "MPECAN_FW_TC_0188",
    "MPECAN_FW_TC_0189",
    "MPECAN_FW_TC_0190",
    "MPECAN_FW_TC_0191",
    "MPECAN_FW_TC_0192",
    "MPECAN_FW_TC_0193",
    "MPECAN_FW_TC_0194",
    "MPECAN_FW_TC_0195",
    "MPECAN_FW_TC_0196",
    "MPECAN_FW_TC_0197",
    "MPECAN_FW_TC_0198",
    "MPECAN_FW_TC_0199",
    "MPECAN_FW_TC_0200"
]
def search_between_strings(file_path, string_a, string_b):
    listTestInfo = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            
            found_a = False
            start_line = 0
            
            for line_num, line in enumerate(lines, 1):
                if string_a in line and not found_a:
                    found_a = True
                    start_line = line_num
                    # print(f"找到字符串A '{string_a}' 在行 {line_num}")
                    continue
                
                if found_a and string_b in line:
                    # print(f"找到字符串B '{string_b}' 在行 {line_num}")
                    # print(f"\n字符串A和字符串B之间的文本（行 {start_line} 到行 {line_num}）:")
                    # 打印从字符串A所在行到字符串B所在行之间的所有文本
                    for i in range(start_line - 1, line_num):
                        if "test begin >>>" in lines[i].strip():
                            for j in range(i - 1, line_num - 1):
                                listTestInfo.append(lines[j].strip())
                    break
                    
    except FileNotFoundError:
        print(f"文件 '{file_path}' 不存在")
    except Exception as e:
        print(f"发生错误: {e}")
    strTestInfo = "\n".join(listTestInfo)
    print(type(strTestInfo))
    return strTestInfo

def get_test_ret_info(tc_act_dict, tc_all_list):
    ls_isTest = []
    ls_log = []
    ls_ret = []

    for i, tc in enumerate(tc_all_list):
        if tc in tc_act_dict.keys():
            ls_isTest.append("Yes")
            ls_log.append(tc_act_dict[tc])
            if "Passed" in tc_act_dict[tc]:
                ls_ret.append("pass")
            else:
                ls_ret.append("fail")
        else:
            ls_isTest.append("No")
            ls_log.append("")
            ls_ret.append("fail")
    print(ls_isTest)
    print(ls_log)
    print(ls_ret)
    return ls_isTest, ls_log, ls_ret

def clean_string_for_excel(text):
    """
    清理字符串，使其适合写入Excel单元格
    """
    if text is None:
        return ""
    
    # 转换为字符串
    text = str(text)
    
    # 移除或替换可能引起问题的特殊字符
    # 移除控制字符（ASCII 0-31，除了制表符、换行符、回车符）
    text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)
    
    # 替换可能引起问题的字符
    text = text.replace('\x00', '')  # 空字符
    text = text.replace('\x01', '')  # 标题开始
    text = text.replace('\x02', '')  # 正文开始
    text = text.replace('\x03', '')  # 正文结束
    text = text.replace('\x04', '')  # 文件结束
    
    # 限制字符串长度（Excel单元格最大字符数为32767）
    if len(text) > 32000:
        text = text[:32000] + "...[内容被截断]"
    
    return text
def write_list_to_column(string_list, excel_file_path, column='P', start_row=2):
    """
    将字符串列表写入Excel文件的指定列
    
    参数:
    string_list: list - 要写入的字符串列表
    excel_file_path: str - Excel文件路径
    column: str - 列字母，默认为'P'
    start_row: int - 起始行号，默认为2（P2列）
    """
    try:
        # 加载工作簿
        wb = load_workbook(filename=excel_file_path)
        
        # 获取第一个工作表（或指定工作表）
        ws = wb["Test Case Report"]
        
        # 将字符串列表写入指定列
        for i, value in enumerate(string_list):
            cleaned_value = clean_string_for_excel(value)
            row_num = start_row + i
            cell_address = f"{column}{row_num}"
            ws[cell_address] = cleaned_value
        
        # 保存工作簿
        wb.save(excel_file_path)
        print(f"成功将{len(string_list)}个字符串写入到{excel_file_path}的{column}列")
        
    except Exception as e:
        print(f"写入Excel文件时出错: {e}")

def searchAllInList(strTcList=case_in_redmine):
    listTcInfos = []
    for strTc in strTcList:
        listTcInfos.append(search_between_strings(report_path, strTc, ">>> test end"))
    return listTcInfos

if __name__ == "__main__":
    # write_list_to_column(list(case_in_redmine.keys()), "C:\\tangyapeng\\docs\\StarGather\\report_gen_proj\\MPECAN-report.xlsx", column='A', start_row=2)
    # write_list_to_column(list(case_in_redmine.values()), "C:\\tangyapeng\\docs\\StarGather\\report_gen_proj\\MPECAN-report.xlsx", column='P', start_row=2)
    ls_isTest, ls_log, ls_ret = get_test_ret_info(case_in_redmine, all_case_s)
    write_list_to_column(ls_isTest, "C:\\tangyapeng\\docs\\StarGather\\report_gen_proj\\MPECAN-report.xlsx", column='O', start_row=2)
    write_list_to_column(ls_log, "C:\\tangyapeng\\docs\\StarGather\\report_gen_proj\\MPECAN-report.xlsx", column='P', start_row=2)
    write_list_to_column(ls_ret, "C:\\tangyapeng\\docs\\StarGather\\report_gen_proj\\MPECAN-report.xlsx", column='Q', start_row=2)

