from openpyxl import load_workbook
from tfcParamMang import ParamsOneCase

def test(fileName = './hmi_doc.xlsx'):
    wb = load_workbook(filename=fileName)
    listSheelName = wb.sheetnames
    print(listSheelName[0], 'type is :', type(listSheelName[0]))


class ParameterLoader:

    '''
    The class ParameterLoader is used to parse a excel file and get parameter.
    '''
    listSheetNames = []
    def __init__(self, excelFileName = './hmi_doc.xlsx'):
        try:
            self.wb = load_workbook(filename=excelFileName)
            self.listSheetNames = self.wb.sheetnames
        except Exception:
            print('Load excel file failed.')
            exit(0)

    # Load parameters of a sheet, and store them into a instance of ParamsOneCase
    # Load parameters and generete scripts in stream, and do not malloc too much memory
    def loadASheet(self, sheetName = None):
        if sheetName == None:
            return
        
        params = ParamsOneCase(sheetName)

        ws = self.wb[sheetName]
        # Get data from current sheet
        for co in range(1,10):
           print(ws.cell(row=1, column=co).value)




if __name__ == '__main__':
    newLoader = ParameterLoader()
    print(newLoader.listSheetNames, 'and the length is:', len(newLoader.listSheetNames), end='\n')
    newLoader.loadASheet(newLoader.listSheetNames[0])
