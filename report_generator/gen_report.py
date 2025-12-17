import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

# ==================== 配置区 ====================
REPORT_FILE = "report.xlsx"
RECORD_FILE = "record.xlsx"

# report.xlsx 中要操作的 sheet 名称
REPORT_SHEET = "Test Case Report"

# record.xlsx 中日志所在的 sheet（默认第一个）
RECORD_SHEET = "Sheet"

# 列映射（根据你的 report.xlsx 实际列位置调整）
# 从 A=1 开始计数
COL_MAPPING = {
    "测试用例ID": 1,      # A列
    "用例发布时间": 13,    # M列（Case Released Date）
    "测试记录":    16,     # P列（Test Log）
    "测试结果":    17,     # Q列（Test Result）
}

# ================================================

def load_record_to_report():
    print("正在加载 report.xlsx ...")
    wb = load_workbook(REPORT_FILE)
    ws_report = wb[REPORT_SHEET]

    print("正在读取 record.xlsx 日志记录...")
    df_record = pd.read_excel(RECORD_FILE, sheet_name=RECORD_SHEET)
    # 确保列名正确（去掉空格）
    df_record.columns = df_record.columns.str.strip()

    # 构建一个字典：{ 部分匹配的 TestCase 名 -> (Time, Log, Result) }
    record_dict = {}
    for _, row in df_record.iterrows():
        testcase_name = str(row['TestCase']).strip()
        time_str = str(row['Time']).strip()
        log_str = str(row['Log']).strip()
        result_str = str(row['Result']).strip()

        # 清理可能的多余字符
        if testcase_name == 'nan':
            continue
        record_dict[testcase_name] = (time_str, log_str, result_str)

    print(f"共加载 {len(record_dict)} 条执行记录，可用于匹配。")

    # 遍历 report 中的每一行（从第2行开始，跳过标题）
    updated_count = 0
    for row_idx in range(2, ws_report.max_row + 1):
        cell_id = ws_report.cell(row=row_idx, column=COL_MAPPING["测试用例ID"]).value
        if not cell_id:
            continue
        testcase_id = str(cell_id).strip()

        # 查找 record 中任意一条 TestCase 名称包含这个 testcase_id 的记录
        matched_key = None
        matched_data = None
        for key in record_dict:
            if testcase_id in key:  # 子串匹配
                matched_key = key
                matched_data = record_dict[key]
                break

        if matched_data:
            time_val, log_val, result_val = matched_data

            # 写入“用例发布时间”（取执行时间）
            ws_report.cell(row=row_idx, column=COL_MAPPING["用例发布时间"]).value = time_val

            # 写入“测试记录”（Log 内容）
            ws_report.cell(row=row_idx, column=COL_MAPPING["测试记录"]).value = log_val

            # 写入“测试结果”（pass / fail）
            clean_result = "Pass" if "pass" in result_val.lower() else "Fail"
            ws_report.cell(row=row_idx, column=COL_MAPPING["测试结果"]).value = clean_result

            updated_count += 1
            print(f"[{updated_count}] {testcase_id} 已填充 -> {time_val} | {clean_result}")

    print(f"\n匹配完成！共更新 {updated_count} 条测试用例结果。")
    wb.save(REPORT_FILE)
    print(f"已保存更新后的文件：{REPORT_FILE}")

if __name__ == "__main__":
    if not os.path.exists(REPORT_FILE):
        print(f"错误：找不到 {REPORT_FILE}")
    elif not os.path.exists(RECORD_FILE):
        print(f"错误：找不到 {RECORD_FILE}")
    else:
        load_record_to_report()
        print("所有操作完成！")